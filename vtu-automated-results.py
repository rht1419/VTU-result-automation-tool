#!/usr/bin/env python3
"""
VTU Results Automation - Web Interface
Ultra-fast web interface for automation
Direct form URL access version
"""

import json
import threading
import time
import re
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO, emit
import os
import sys

# Import PDF extraction functionality
try:
    import PyPDF2
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    PDF_EXTRACTION_AVAILABLE = True
    print("✅ PDF extraction libraries available")
except ImportError as e:
    PDF_EXTRACTION_AVAILABLE = False
    print(f"⚠️ PDF extraction libraries not available: {e}")
    print("Install with: pip install PyPDF2 pandas openpyxl")

# Import the ultra-fast automation class
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
try:
    from latest_vtu_automation_ultra_fast import VTUAutomationUltraFast  # Use ultra-fast version
except ImportError:
    print("❌ Could not import VTUAutomationUltraFast. Please ensure 'latest_vtu_automation_ultra_fast.py' exists in the same directory.")
    sys.exit(1)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'vtu_automation_ultra_fast_2024'
socketio = SocketIO(app, cors_allowed_origins="*")

# Global variables for automation status
automation_status = {
    'is_running': False,
    'current_usn': '',
    'progress': 0,
    'total': 0,
    'successful': 0,
    'failed': 0,
    'current_batch': None,
    'logs': []
}

class WebSocketLogger:
    """Custom logger that emits to WebSocket"""
    def __init__(self):
        self.logs = []
    
    def log(self, message, level='info'):
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_entry = {
            'timestamp': timestamp,
            'message': message,
            'level': level
        }
        self.logs.append(log_entry)
        automation_status['logs'] = self.logs[-50:]  # Keep last 50 logs
        socketio.emit('log_update', log_entry)

logger = WebSocketLogger()

class UltraFastWebAutomation(VTUAutomationUltraFast):
    """Ultra-fast automation class with optimized settings"""
    
    def __init__(self, headless=True, enable_pdf=False):
        super().__init__(headless, enable_pdf)
        self.logger = logger
        
        # ULTRA-FAST OPTIMIZATIONS
        self.max_captcha_retries = 5
        self.page_timeout = 8000       # Reduced from 15s → 8s
        self.request_delay = 0.1       # Ultra-minimal delay
        self.element_timeout = 4000    # Faster element detection
        self.result_timeout = 2500     # Faster result detection
    
    def process_batch_with_updates(self, usns, form_url, batch_name="Ultra-Fast Batch"):
        """Process batch with real-time WebSocket updates and ultra-fast settings"""
        global automation_status
        
        # Ensure PDF files are saved under this batch's subfolder
        try:
            self.current_batch_name = batch_name
        except Exception:
            pass

        automation_status.update({
            'is_running': True,
            'current_batch': batch_name,
            'total': len(usns),
            'progress': 0,
            'successful': 0,
            'failed': 0
        })
        
        self.logger.log(f"🚀 Starting ULTRA-FAST batch: {batch_name} ({len(usns)} USNs)", 'success')
        self.logger.log(f"🔗 Using form URL: {form_url}", 'info')
        socketio.emit('status_update', automation_status)
        
        batch_result = {
            'batch_name': batch_name,
            'form_url': form_url,
            'total': len(usns),
            'successful': 0,
            'failed': 0,
            'results': [],
            'start_time': datetime.now().isoformat()
        }
        
        for i, usn in enumerate(usns, 1):
            if not automation_status['is_running']:
                self.logger.log("❌ Ultra-fast batch stopped by user", 'warning')
                break
                
            automation_status['current_usn'] = usn
            automation_status['progress'] = i
            
            self.logger.log(f"⚡ Ultra-fast processing {i}/{len(usns)}: {usn}", 'info')
            socketio.emit('status_update', automation_status)
            
            result = self.process_single_usn(usn, form_url)
            batch_result['results'].append(result)
            
            if result['success']:
                batch_result['successful'] += 1
                automation_status['successful'] += 1
                self.logger.log(f"✅ {usn}: Ultra-fast success - Files saved", 'success')
            else:
                batch_result['failed'] += 1
                automation_status['failed'] += 1
                self.logger.log(f"❌ {usn}: {result['error']}", 'error')
            
            socketio.emit('status_update', automation_status)
            
            # Ultra-minimal delay between requests
            if i < len(usns) and automation_status['is_running']:
                self.logger.log(f"⚡ Ultra-fast delay: {self.request_delay:.1f}s...", 'info')
                time.sleep(self.request_delay)
        
        batch_result['end_time'] = datetime.now().isoformat()
        
        # Save batch summary
        summary_path = self.results_dir / f"summary_ultrafast_{batch_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(summary_path, 'w') as f:
            json.dump(batch_result, f, indent=2)
        
        success_rate = (batch_result['successful'] / batch_result['total'] * 100) if batch_result['total'] > 0 else 0
        
        automation_status['is_running'] = False
        automation_status['current_usn'] = ''
        
        self.logger.log(f"🎉 ULTRA-FAST batch completed! Success rate: {success_rate:.1f}%", 'success')
        self.logger.log(f"✅ Successful: {batch_result['successful']} | ❌ Failed: {batch_result['failed']}", 'info')
        self.logger.log(f"📁 Summary saved: {summary_path.name}", 'info')
        
        socketio.emit('status_update', automation_status)
        socketio.emit('batch_complete', {
            'success_rate': success_rate,
            'summary_file': summary_path.name,
            'batch_result': batch_result
        })
        
        return batch_result

@app.route('/')
def index():
    return render_template('vtu_interface.html')

@app.route('/api/start_batch', methods=['POST'])
def start_batch():
    global automation_status
    
    if automation_status['is_running']:
        return jsonify({'success': False, 'message': 'Ultra-fast automation is already running'})
    
    # Get JSON data from request
    data = request.get_json()
    if data is None:
        return jsonify({'success': False, 'message': 'Invalid request data'})
    
    # Extract parameters with defaults
    usns_data = data.get('usns', [])
    usns = [usn.strip().upper() for usn in usns_data if usn.strip()] if isinstance(usns_data, list) else []
    form_url = data.get('form_url', '').strip() if data.get('form_url') else ''
    batch_name = data.get('batch_name', 'Ultra-Fast Batch')
    headless_mode = data.get('headless', True)
    enable_pdf = data.get('enable_pdf', True)  # Enable PDF by default
    
    if not usns:
        return jsonify({'success': False, 'message': 'No valid USNs provided'})
    
    if not form_url:
        return jsonify({'success': False, 'message': 'Form URL is required'})
    
    # Validate form URL
    if not form_url.startswith('http'):
        return jsonify({'success': False, 'message': 'Invalid form URL. Must start with http:// or https://'})
    
    # Start ultra-fast automation in background thread
    def run_ultra_fast_automation():
        try:
            automation = UltraFastWebAutomation(headless=headless_mode, enable_pdf=enable_pdf)
            automation.process_batch_with_updates(usns, form_url, batch_name)
        except Exception as e:
            logger.log(f"❌ Ultra-fast automation error: {str(e)}", 'error')
            automation_status['is_running'] = False
            socketio.emit('status_update', automation_status)
    
    thread = threading.Thread(target=run_ultra_fast_automation, daemon=True)
    thread.start()
    
    return jsonify({'success': True, 'message': f'Started ultra-fast processing of {len(usns)} USNs'})

@app.route('/api/stop_batch', methods=['POST'])
def stop_batch():
    global automation_status
    automation_status['is_running'] = False
    logger.log("🛑 Ultra-fast processing stop requested by user", 'warning')
    return jsonify({'success': True, 'message': 'Ultra-fast stop signal sent'})

@app.route('/api/status')
def get_status():
    return jsonify(automation_status)

@app.route('/api/results')
def list_results():
    pdf_dir = Path('pdf_results')
    if not pdf_dir.exists():
        return jsonify([])

    files = []
    for file_path in pdf_dir.rglob('*.pdf'):
        if file_path.is_file():
            rel_path = file_path.relative_to(pdf_dir).as_posix()
            files.append({
                'name': file_path.name,
                'path': rel_path,
                'size': file_path.stat().st_size,
                'modified': datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()
            })

    return jsonify(sorted(files, key=lambda x: x['modified'], reverse=True))


@app.route('/api/captcha_status')
def captcha_status():
    """
    Check if there are captcha images and return count
    """
    try:
        captcha_dir = Path('captchas').resolve()
        if not captcha_dir.exists():
            return jsonify({'has_captchas': False, 'count': 0})
        
        captcha_count = 0
        for captcha_file in captcha_dir.rglob('*'):
            if captcha_file.is_file() and (captcha_file.suffix.lower() in ['.png', '.jpg', '.jpeg']):
                captcha_count += 1
        
        return jsonify({
            'has_captchas': captcha_count > 0,
            'count': captcha_count
        })
    except Exception as e:
        logger.log(f"❌ Error checking captcha status: {e}", 'error')
        return jsonify({'has_captchas': False, 'count': 0})

@app.route('/api/download/<path:filename>')
def download_file(filename):
    pdf_dir = Path('pdf_results').resolve()
    requested_path = (pdf_dir / filename).resolve()

    # Security: ensure requested path is inside pdf_dir
    if not str(requested_path).startswith(str(pdf_dir)):
        return jsonify({'error': 'Invalid path'}), 400

    if requested_path.exists() and requested_path.is_file():
        return send_file(str(requested_path), as_attachment=True)
    else:
        return jsonify({'error': 'File not found'}), 404


@app.route('/api/delete/<path:filename>', methods=['DELETE'])
def delete_file(filename):
    pdf_dir = Path('pdf_results').resolve()
    requested_path = (pdf_dir / filename).resolve()
    
    # Security: ensure requested path is inside pdf_dir
    if not str(requested_path).startswith(str(pdf_dir)):
        return jsonify({'success': False, 'message': 'Invalid path'}), 400
    
    if requested_path.exists() and requested_path.is_file():
        try:
            requested_path.unlink()
            logger.log(f"🗑️ Deleted file: {filename}", 'info')
            return jsonify({'success': True, 'message': 'File deleted successfully'})
        except Exception as e:
            logger.log(f"❌ Error deleting file {filename}: {e}", 'error')
            return jsonify({'success': False, 'message': f'Error deleting file: {str(e)}'}), 500
    else:
        return jsonify({'success': False, 'message': 'File not found'}), 404


@app.route('/api/delete_captchas', methods=['DELETE'])
def delete_captchas():
    """
    Delete all captcha images to free up space
    """
    try:
        captcha_dir = Path('captchas').resolve()
        if not captcha_dir.exists():
            return jsonify({'success': True, 'message': 'Captcha directory does not exist'})
        
        deleted_count = 0
        for captcha_file in captcha_dir.rglob('*'):
            if captcha_file.is_file() and (captcha_file.suffix.lower() in ['.png', '.jpg', '.jpeg']):
                captcha_file.unlink()
                deleted_count += 1
        
        logger.log(f"🗑️ Deleted {deleted_count} captcha images", 'info')
        return jsonify({
            'success': True, 
            'message': f'Successfully deleted {deleted_count} captcha images',
            'deleted_count': deleted_count
        })
    except Exception as e:
        logger.log(f"❌ Error deleting captcha images: {e}", 'error')
        return jsonify({
            'success': False, 
            'message': f'Error deleting captcha images: {str(e)}'
        }), 500


@app.route('/api/delete_all_pdfs', methods=['DELETE'])
def delete_all_pdfs():
    """
    Delete all PDF files from the pdf_results directory
    """
    try:
        pdf_dir = Path('pdf_results').resolve()
        if not pdf_dir.exists():
            return jsonify({'success': True, 'message': 'PDF directory does not exist'})
        
        deleted_count = 0
        for pdf_file in pdf_dir.rglob('*.pdf'):
            if pdf_file.is_file():
                pdf_file.unlink()
                deleted_count += 1
        
        logger.log(f"🗑️ Deleted {deleted_count} PDF files", 'info')
        return jsonify({
            'success': True, 
            'message': f'Successfully deleted {deleted_count} PDF files',
            'deleted_count': deleted_count
        })
    except Exception as e:
        logger.log(f"❌ Error deleting PDF files: {e}", 'error')
        return jsonify({
            'success': False, 
            'message': f'Error deleting PDF files: {str(e)}'
        }), 500

@socketio.on('connect')
def handle_connect():
    emit('status_update', automation_status)
    emit('log_update', {'logs': automation_status['logs']})


# PDF extraction and conversion functions

def extract_text_from_pdf(pdf_path):
    """
    Extract text from a PDF file
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        str: Extracted text from the PDF
    """
    if not PDF_EXTRACTION_AVAILABLE:
        logger.log("❌ PDF extraction libraries not available", 'error')
        return ""
    
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        logger.log(f"❌ Error reading PDF {pdf_path}: {e}", 'error')
        return ""

def parse_vtu_result_text(text):
    """
    Parse VTU result text and extract relevant information using advanced parsing logic
    
    Args:
        text (str): Text extracted from VTU result PDF
        
    Returns:
        dict: Parsed result data
    """
    # Initialize result data structure
    result_data = {
        'usn': '',
        'student_name': '',
        'semester': '',
        'result': '',
        'sgpa': '',
        'cgpa': '',
        'subjects': [],
        'total_marks': '',
        'credits_earned': '',
        'credits_registered': ''
    }
    
    # Extract USN
    usn_match = re.search(r'University Seat Number\s*:\s*([A-Z0-9]+)', text, re.IGNORECASE)
    if usn_match:
        result_data['usn'] = usn_match.group(1)
    
    # Extract Student Name
    name_match = re.search(r'Student Name\s*:\s*([A-Z\s]+)', text, re.IGNORECASE)
    if name_match:
        result_data['student_name'] = name_match.group(1).strip()
    
    # Extract Semester
    sem_match = re.search(r'Semester\s*:\s*(\d+)', text, re.IGNORECASE)
    if sem_match:
        result_data['semester'] = sem_match.group(1)
    
    # Extract SGPA
    sgpa_match = re.search(r'SGPA\s*:\s*([\d.]+)', text, re.IGNORECASE)
    if sgpa_match:
        result_data['sgpa'] = sgpa_match.group(1)
    
    # Extract CGPA
    cgpa_match = re.search(r'CGPA\s*:\s*([\d.]+)', text, re.IGNORECASE)
    if cgpa_match:
        result_data['cgpa'] = cgpa_match.group(1)
    
    # Extract Result
    result_match = re.search(r'Result\s*:\s*(PASS|FAIL)', text, re.IGNORECASE)
    if result_match:
        result_data['result'] = result_match.group(1).upper()
    
    # Extract subject data
    # Handle the multi-line format in VTU PDFs
    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Look for subject entries that have the code at the beginning of the line
        # Pattern: Subject code followed by subject name on the same line
        # Example: BCS401 ANAL YSIS & DESIGN OF
        # Then the next line has: ALGORITHMS48 34 82 P 2025-07-31
        
        # Try to match subject code at the beginning of the line
        subject_line_match = re.match(r'^([A-Z0-9]+)\s+(.+)$', line)
        if subject_line_match:
            subject_code = subject_line_match.group(1)
            subject_name_part = subject_line_match.group(2)
            
            # Check if next line contains the marks or continuation of subject name
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                
                # Check if next line has marks (ends with numbers and P/F)
                subject_pattern = r'(.+?)(\d{1,2})\s+(\d{1,2})\s+(\d{1,3})\s+([PF])'
                match = re.search(subject_pattern, next_line)
                
                if match:
                    # The subject name is split across multiple lines
                    # First part from current line, additional parts from intermediate lines
                    marks_name = match.group(1).strip()
                    internal = match.group(2)
                    external = match.group(3)
                    total = match.group(4)
                    result = match.group(5).upper()
                    
                    # Start with the subject name part from the code line
                    full_subject_name = subject_name_part
                    
                    # Check for additional subject name lines between code line and marks line
                    # Look for intermediate lines that are just words (no numbers at the end)
                    j = i + 1
                    while j > i + 1 and j < len(lines) and not re.search(r'\d+\s+\d+\s+\d+\s+[PF]', lines[j].strip()):
                        if lines[j].strip() and not re.search(r'\d+$', lines[j].strip()):
                            full_subject_name += ' ' + lines[j].strip()
                        j -= 1
                    
                    # Add the marks line subject name part
                    full_subject_name += ' ' + marks_name
                    
                    subject_data = {
                        'subject_code': subject_code,
                        'subject_name': full_subject_name.strip(),
                        'internal': internal,
                        'external': external,
                        'total': total,
                        'result': result
                    }
                    result_data['subjects'].append(subject_data)
                    i += 2  # Skip the next line as we've processed it
                    continue
                else:
                    # Check if this might be a multi-line subject name
                    # If next line doesn't have marks, it might be continuation of subject name
                    if i + 2 < len(lines):
                        third_line = lines[i + 2].strip()
                        third_match = re.search(subject_pattern, third_line)
                        if third_match:
                            # Three-line subject: code line -> name continuation line -> marks line
                            name_continuation = next_line
                            marks_name = third_match.group(1).strip()
                            internal = third_match.group(2)
                            external = third_match.group(3)
                            total = third_match.group(4)
                            result = third_match.group(5).upper()
                            
                            # Combine all parts
                            full_subject_name = subject_name_part + ' ' + name_continuation + ' ' + marks_name
                            
                            subject_data = {
                                'subject_code': subject_code,
                                'subject_name': full_subject_name.strip(),
                                'internal': internal,
                                'external': external,
                                'total': total,
                                'result': result
                            }
                            result_data['subjects'].append(subject_data)
                            i += 3  # Skip the next two lines as we've processed them
                            continue
        
        # Also try a simpler pattern for cases where all data is on one line
        simple_pattern = r'([A-Z0-9]+)\s+(.+?)(\d{1,2})\s+(\d{1,2})\s+(\d{1,3})\s+([PF])'
        simple_match = re.search(simple_pattern, line)
        if simple_match:
            subject_data = {
                'subject_code': simple_match.group(1),
                'subject_name': simple_match.group(2).strip(),
                'internal': simple_match.group(3),
                'external': simple_match.group(4),
                'total': simple_match.group(5),
                'result': simple_match.group(6).upper()
            }
            result_data['subjects'].append(subject_data)
        
        i += 1
    
    # Extract total marks
    total_match = re.search(r'Total Marks\s*:\s*(\d+)', text, re.IGNORECASE)
    if total_match:
        result_data['total_marks'] = total_match.group(1)
    
    # Extract credits
    credits_match = re.search(r'Credits\s*:\s*(\d+)\s*/\s*(\d+)', text, re.IGNORECASE)
    if credits_match:
        result_data['credits_registered'] = credits_match.group(1)
        result_data['credits_earned'] = credits_match.group(2)
    
    return result_data

def process_single_pdf(pdf_path):
    """
    Process a single PDF file and extract result data
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        dict: Extracted result data or None if failed
    """
    logger.log(f"📄 Processing PDF: {os.path.basename(pdf_path)}", 'info')
    
    # Extract text from PDF
    text = extract_text_from_pdf(pdf_path)
    if not text:
        return None
    
    # Parse the extracted text
    result_data = parse_vtu_result_text(text)
    
    # Add file information
    result_data['pdf_file'] = os.path.basename(pdf_path)
    result_data['pdf_path'] = str(pdf_path)
    
    logger.log(f"  ✅ Extracted data for USN: {result_data['usn']}", 'success')
    return result_data

def process_pdf_batch(pdf_directory):
    """
    Process PDF files from the latest batch in the directory
    
    Args:
        pdf_directory (str): Path to directory containing PDF files
        
    Returns:
        list: List of extracted result data
    """
    if not PDF_EXTRACTION_AVAILABLE:
        logger.log("❌ PDF extraction libraries not available", 'error')
        return []
    
    pdf_directory = Path(pdf_directory)
    results = []
    
    if not pdf_directory.exists():
        logger.log(f"❌ Directory not found: {pdf_directory}", 'error')
        return results
    
    # Find all batch directories
    batch_dirs = [d for d in pdf_directory.iterdir() if d.is_dir()]
    
    if not batch_dirs:
        logger.log("⚠️ No batch directories found, checking for PDF files in root directory", 'warning')
        # Check for PDF files directly in the root directory
        pdf_files = list(pdf_directory.rglob("*.pdf"))
        pdf_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    else:
        # Sort batch directories by modification time to get the latest one
        batch_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        latest_batch = batch_dirs[0]
        logger.log(f"📂 Latest batch directory: {latest_batch.name}", 'info')
        
        # Find all PDF files in the latest batch directory
        pdf_files = list(latest_batch.rglob("*.pdf"))
        pdf_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    
    logger.log(f"🔍 Found {len(pdf_files)} PDF files to process (sorted by newest first)", 'info')
    
    # Limit to the 100 most recent files to avoid processing old files
    pdf_files = pdf_files[:100]
    
    # Process the files
    for i, pdf_file in enumerate(pdf_files):
        # Log the first few files as newest
        if i < 10:
            logger.log(f"📄 Processing file {i+1}/{len(pdf_files)}: {pdf_file.name} (Modified: {datetime.fromtimestamp(pdf_file.stat().st_mtime)})", 'info')
        result_data = process_single_pdf(pdf_file)
        if result_data:
            results.append(result_data)
    
    return results

def create_summary_excel(results, output_path):
    """
    Create a summary Excel file with all results
    
    Args:
        results (list): List of result data dictionaries
        output_path (str): Path to output Excel file
    """
    if not PDF_EXTRACTION_AVAILABLE:
        logger.log("❌ PDF extraction libraries not available", 'error')
        return
        
    if not results:
        logger.log("❌ No results to export", 'error')
        return
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "VTU Results Summary"
    
    # Define headers for summary sheet
    headers = [
        'PDF File', 'USN', 'Student Name', 'Semester', 'Result', 
        'SGPA', 'CGPA', 'Total Marks', 'Credits Registered', 'Credits Earned'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write summary data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('pdf_file', ''))
        ws.cell(row=row, column=2, value=result.get('usn', ''))
        ws.cell(row=row, column=3, value=result.get('student_name', ''))
        ws.cell(row=row, column=4, value=result.get('semester', ''))
        ws.cell(row=row, column=5, value=result.get('result', ''))
        ws.cell(row=row, column=6, value=result.get('sgpa', ''))
        ws.cell(row=row, column=7, value=result.get('cgpa', ''))
        ws.cell(row=row, column=8, value=result.get('total_marks', ''))
        ws.cell(row=row, column=9, value=result.get('credits_registered', ''))
        ws.cell(row=row, column=10, value=result.get('credits_earned', ''))
    
    # Save workbook (only summary sheet, no subject details)
    try:
        wb.save(output_path)
        logger.log(f"✅ Summary Excel file saved: {output_path}", 'success')
    except Exception as e:
        logger.log(f"❌ Error saving Excel file: {e}", 'error')

@app.route('/api/convert_to_excel', methods=['POST'])
def convert_to_excel():
    """
    Convert all PDF results to Excel format using advanced subject-wise DataFrame
    """
    if not PDF_EXTRACTION_AVAILABLE:
        return jsonify({'success': False, 'message': 'PDF extraction libraries not available. Install with: pip install PyPDF2 pandas openpyxl'})
    
    try:
        # Get the project root directory
        project_root = Path(__file__).parent
        pdf_results_dir = project_root / "pdf_results"
        output_dir = project_root / "excel_results"
        output_dir.mkdir(exist_ok=True)
        
        logger.log("🚀 Starting PDF to Excel conversion...", 'info')
        logger.log(f"📂 PDF Results Directory: {pdf_results_dir}", 'info')
        logger.log(f"📂 Output Directory: {output_dir}", 'info')
        
        # Process all PDF files
        results = process_pdf_batch(pdf_results_dir)
        
        if not results:
            return jsonify({'success': False, 'message': 'No PDF results found to process'})
        
        logger.log(f"✅ Processed {len(results)} PDF files", 'success')
        
        # Log some details about the results
        if results:
            logger.log(f"📊 First result USN: {results[0].get('usn', 'N/A')}", 'info')
            logger.log(f"📊 Last result USN: {results[-1].get('usn', 'N/A')}", 'info')
        
        # Create subject-wise DataFrame with the exact format requested
        # Collect all unique subject codes (not names)
        all_subjects = set()
        for result in results:
            for subject in result.get('subjects', []):
                all_subjects.add(subject.get('subject_code', '').strip())
        
        # Sort subject codes for consistent column ordering
        subject_columns = sorted(list(all_subjects))
        logger.log(f"📚 Found {len(subject_columns)} unique subject codes: {', '.join(subject_columns)}", 'info')
        
        # Create rows for each student
        rows = []
        for result in results:
            # Clean up the student name
            student_name = result.get('student_name', '')
            if '\n' in student_name:
                student_name = student_name.split('\n')[0].strip()
            
            row = {
                'USN': result.get('usn', ''),
                'Name': student_name,  # Changed from 'Student Name' to 'Name'
            }
            
            # Initialize total marks counter
            total_marks = 0
            
            # Add subject marks
            for subject_code in subject_columns:
                row[subject_code] = ''  # Initialize with empty string
            
            # Fill in actual subject marks
            for subject in result.get('subjects', []):
                subject_code = subject.get('subject_code', '').strip()
                if subject_code in row:
                    subject_total = subject.get('total', '')
                    row[subject_code] = subject_total
                    # Add to total marks if it's a valid number
                    if subject_total.isdigit():
                        total_marks += int(subject_total)
            
            # Add CGPA and SGPA from result data
            row['CGPA'] = result.get('cgpa', '')
            row['SGPA'] = result.get('sgpa', '')
            
            # Add total marks to the row
            row['Total Marks'] = str(total_marks)
            
            # Determine overall result
            subject_results = [subject.get('result', '').upper() for subject in result.get('subjects', [])]
            if all(res == 'P' for res in subject_results) and subject_results:
                row['Result'] = 'P'  # Changed from 'RESULT' to 'Result'
            elif any(res == 'F' for res in subject_results):
                # Find the first failed subject
                failed_subject = None
                for subject in result.get('subjects', []):
                    if subject.get('result', '').upper() == 'F':
                        failed_subject = subject.get('subject_code', '').strip()
                        break
                row['Result'] = f'F in {failed_subject}' if failed_subject else 'F'  # Changed from 'RESULT' to 'Result'
            else:
                row['Result'] = 'N/A'  # Changed from 'RESULT' to 'Result'
            
            rows.append(row)
        
        # Define column order to match exact requirements: USN, Name, subject codes, Total Marks, Result, CGPA, SGPA
        columns = ['USN', 'Name'] + subject_columns + ['Total Marks', 'Result', 'CGPA', 'SGPA']
        
        # Create DataFrame with the exact column structure
        subject_wise_df = pd.DataFrame(rows, columns=columns)
        
        # Create Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = output_dir / f"vtu_results_{timestamp}.xlsx"
        subject_wise_df.to_excel(excel_path, index=False)
        logger.log(f"✅ Excel file saved: {excel_path}", 'success')
        
        logger.log("🎉 PDF to Excel conversion completed successfully!", 'success')
        
        return jsonify({
            'success': True, 
            'message': f'Conversion completed successfully! Generated {len(results)} records.',
            'files': {
                'excel': str(excel_path.relative_to(project_root))
            }
        })
        
    except Exception as e:
        logger.log(f"❌ Error during PDF to Excel conversion: {e}", 'error')
        import traceback
        logger.log(f"❌ Traceback: {traceback.format_exc()}", 'error')
        return jsonify({'success': False, 'message': f'Error during conversion: {str(e)}'})

if __name__ == '__main__':
    # Ensure results directory exists
    Path('results').mkdir(exist_ok=True)
    Path('captchas').mkdir(exist_ok=True)
    Path('pdf_results').mkdir(exist_ok=True)
    
    print("🚀 Starting VTU Automation Web Interface...")
    print("⚡ Maximum speed optimization enabled!")
    print("📱 Access at: http://localhost:5000")
    
    socketio.run(app, host='localhost', port=5000, debug=False)