#!/usr/bin/env python3
"""
PDF to Excel Extractor for VTU Results
Extracts data from VTU result PDFs and converts them to Excel format
"""

import os
import sys
import re
import json
from pathlib import Path
from collections import defaultdict

# Import required libraries
try:
    import PyPDF2
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    print("✅ Required libraries imported successfully")
except ImportError as e:
    print(f"❌ Error importing required libraries: {e}")
    print("Please install required packages: pip install PyPDF2 pandas openpyxl")
    sys.exit(1)

def extract_text_from_pdf(pdf_path):
    """
    Extract text from a PDF file
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        str: Extracted text from the PDF
    """
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        print(f"❌ Error reading PDF {pdf_path}: {e}")
        return ""

def parse_vtu_result_text(text):
    """
    Parse VTU result text and extract relevant information
    
    Args:
        text (str): Text extracted from VTU result PDF
        
    Returns:
        dict: Parsed result data
    """
    # Initialize result data structure
    result_data = {
        'usn': '',
        'student_name': '',
        'semester': '',
        'result': '',
        'sgpa': '',
        'cgpa': '',
        'subjects': [],
        'total_marks': '',
        'credits_earned': '',
        'credits_registered': ''
    }
    
    # Extract USN
    usn_match = re.search(r'University Seat Number\s*:\s*([A-Z0-9]+)', text, re.IGNORECASE)
    if usn_match:
        result_data['usn'] = usn_match.group(1)
    
    # Extract Student Name
    name_match = re.search(r'Student Name\s*:\s*([A-Z\s]+)', text, re.IGNORECASE)
    if name_match:
        result_data['student_name'] = name_match.group(1).strip()
    
    # Extract Semester
    sem_match = re.search(r'Semester\s*:\s*(\d+)', text, re.IGNORECASE)
    if sem_match:
        result_data['semester'] = sem_match.group(1)
    
    # Extract SGPA
    sgpa_match = re.search(r'SGPA\s*:\s*([\d.]+)', text, re.IGNORECASE)
    if sgpa_match:
        result_data['sgpa'] = sgpa_match.group(1)
    
    # Extract CGPA
    cgpa_match = re.search(r'CGPA\s*:\s*([\d.]+)', text, re.IGNORECASE)
    if cgpa_match:
        result_data['cgpa'] = cgpa_match.group(1)
    
    # Extract Result
    result_match = re.search(r'Result\s*:\s*(PASS|FAIL)', text, re.IGNORECASE)
    if result_match:
        result_data['result'] = result_match.group(1).upper()
    
    # Extract subject data
    # Handle the multi-line format in VTU PDFs
    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Look for subject entries that have the code at the beginning of the line
        # Pattern: Subject code followed by subject name on the same line
        # Example: BCS401 ANAL YSIS & DESIGN OF
        # Then the next line has: ALGORITHMS48 34 82 P 2025-07-31
        
        # Try to match subject code at the beginning of the line
        subject_line_match = re.match(r'^([A-Z0-9]+)\s+(.+)$', line)
        if subject_line_match:
            subject_code = subject_line_match.group(1)
            subject_name_part = subject_line_match.group(2)
            
            # Check if next line contains the marks or continuation of subject name
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                
                # Check if next line has marks (ends with numbers and P/F)
                subject_pattern = r'(.+?)(\d{1,2})\s+(\d{1,2})\s+(\d{1,3})\s+([PF])'
                match = re.search(subject_pattern, next_line)
                
                if match:
                    # The subject name is split across multiple lines
                    # First part from current line, additional parts from intermediate lines
                    marks_name = match.group(1).strip()
                    internal = match.group(2)
                    external = match.group(3)
                    total = match.group(4)
                    result = match.group(5).upper()
                    
                    # Start with the subject name part from the code line
                    full_subject_name = subject_name_part
                    
                    # Check for additional subject name lines between code line and marks line
                    # Look for intermediate lines that are just words (no numbers at the end)
                    j = i + 1
                    while j > i + 1 and j < len(lines) and not re.search(r'\d+\s+\d+\s+\d+\s+[PF]', lines[j].strip()):
                        if lines[j].strip() and not re.search(r'\d+$', lines[j].strip()):
                            full_subject_name += ' ' + lines[j].strip()
                        j -= 1
                    
                    # Add the marks line subject name part
                    full_subject_name += ' ' + marks_name
                    
                    subject_data = {
                        'subject_code': subject_code,
                        'subject_name': full_subject_name.strip(),
                        'internal': internal,
                        'external': external,
                        'total': total,
                        'result': result
                    }
                    result_data['subjects'].append(subject_data)
                    i += 2  # Skip the next line as we've processed it
                    continue
                else:
                    # Check if this might be a multi-line subject name
                    # If next line doesn't have marks, it might be continuation of subject name
                    if i + 2 < len(lines):
                        third_line = lines[i + 2].strip()
                        third_match = re.search(subject_pattern, third_line)
                        if third_match:
                            # Three-line subject: code line -> name continuation line -> marks line
                            name_continuation = next_line
                            marks_name = third_match.group(1).strip()
                            internal = third_match.group(2)
                            external = third_match.group(3)
                            total = third_match.group(4)
                            result = third_match.group(5).upper()
                            
                            # Combine all parts
                            full_subject_name = subject_name_part + ' ' + name_continuation + ' ' + marks_name
                            
                            subject_data = {
                                'subject_code': subject_code,
                                'subject_name': full_subject_name.strip(),
                                'internal': internal,
                                'external': external,
                                'total': total,
                                'result': result
                            }
                            result_data['subjects'].append(subject_data)
                            i += 3  # Skip the next two lines as we've processed them
                            continue
        
        # Also try a simpler pattern for cases where all data is on one line
        simple_pattern = r'([A-Z0-9]+)\s+(.+?)(\d{1,2})\s+(\d{1,2})\s+(\d{1,3})\s+([PF])'
        simple_match = re.search(simple_pattern, line)
        if simple_match:
            subject_data = {
                'subject_code': simple_match.group(1),
                'subject_name': simple_match.group(2).strip(),
                'internal': simple_match.group(3),
                'external': simple_match.group(4),
                'total': simple_match.group(5),
                'result': simple_match.group(6).upper()
            }
            result_data['subjects'].append(subject_data)
        
        i += 1
    
    # Extract total marks
    total_match = re.search(r'Total Marks\s*:\s*(\d+)', text, re.IGNORECASE)
    if total_match:
        result_data['total_marks'] = total_match.group(1)
    
    # Extract credits
    credits_match = re.search(r'Credits\s*:\s*(\d+)\s*/\s*(\d+)', text, re.IGNORECASE)
    if credits_match:
        result_data['credits_registered'] = credits_match.group(1)
        result_data['credits_earned'] = credits_match.group(2)
    
    return result_data

def process_single_pdf(pdf_path):
    """
    Process a single PDF file and extract result data
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        dict: Extracted result data or None if failed
    """
    print(f"📄 Processing PDF: {os.path.basename(pdf_path)}")
    
    # Extract text from PDF
    text = extract_text_from_pdf(pdf_path)
    if not text:
        return None
    
    # Parse the extracted text
    result_data = parse_vtu_result_text(text)
    
    # Add file information
    result_data['pdf_file'] = os.path.basename(pdf_path)
    result_data['pdf_path'] = str(pdf_path)
    
    print(f"  ✅ Extracted data for USN: {result_data['usn']}")
    return result_data

def process_pdf_batch(pdf_directory):
    """
    Process all PDF files in a directory and subdirectories
    
    Args:
        pdf_directory (str): Path to directory containing PDF files
        
    Returns:
        list: List of extracted result data
    """
    pdf_directory = Path(pdf_directory)
    results = []
    
    if not pdf_directory.exists():
        print(f"❌ Directory not found: {pdf_directory}")
        return results
    
    # Find all PDF files in the directory and subdirectories
    pdf_files = list(pdf_directory.rglob("*.pdf"))
    print(f"🔍 Found {len(pdf_files)} PDF files to process")
    
    for pdf_file in pdf_files:
        result_data = process_single_pdf(pdf_file)
        if result_data:
            results.append(result_data)
    
    return results

def create_summary_excel(results, output_path):
    """
    Create a summary Excel file with all results
    
    Args:
        results (list): List of result data dictionaries
        output_path (str): Path to output Excel file
    """
    if not results:
        print("❌ No results to export")
        return
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "VTU Results Summary"
    
    # Define headers for summary sheet
    headers = [
        'PDF File', 'USN', 'Student Name', 'Semester', 'Result', 
        'SGPA', 'CGPA', 'Total Marks', 'Credits Registered', 'Credits Earned'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write summary data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('pdf_file', ''))
        ws.cell(row=row, column=2, value=result.get('usn', ''))
        ws.cell(row=row, column=3, value=result.get('student_name', ''))
        ws.cell(row=row, column=4, value=result.get('semester', ''))
        ws.cell(row=row, column=5, value=result.get('result', ''))
        ws.cell(row=row, column=6, value=result.get('sgpa', ''))
        ws.cell(row=row, column=7, value=result.get('cgpa', ''))
        ws.cell(row=row, column=8, value=result.get('total_marks', ''))
        ws.cell(row=row, column=9, value=result.get('credits_registered', ''))
        ws.cell(row=row, column=10, value=result.get('credits_earned', ''))
    
    # Create detailed subject sheet
    if any(result.get('subjects') for result in results):
        ws2 = wb.create_sheet("Subject Details")
        
        # Define headers for subject details
        subject_headers = [
            'PDF File', 'USN', 'Student Name', 'Subject Code', 'Subject Name',
            'Internal', 'External', 'Total', 'Result'
        ]
        
        # Write headers
        for col, header in enumerate(subject_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write subject data
        row = 2
        for result in results:
            usn = result.get('usn', '')
            student_name = result.get('student_name', '')
            pdf_file = result.get('pdf_file', '')
            
            for subject in result.get('subjects', []):
                ws2.cell(row=row, column=1, value=pdf_file)
                ws2.cell(row=row, column=2, value=usn)
                ws2.cell(row=row, column=3, value=student_name)
                ws2.cell(row=row, column=4, value=subject.get('subject_code', ''))
                ws2.cell(row=row, column=5, value=subject.get('subject_name', ''))
                ws2.cell(row=row, column=6, value=subject.get('internal', ''))
                ws2.cell(row=row, column=7, value=subject.get('external', ''))
                ws2.cell(row=row, column=8, value=subject.get('total', ''))
                ws2.cell(row=row, column=9, value=subject.get('result', ''))
                row += 1
    
    # Save workbook
    try:
        wb.save(output_path)
        print(f"✅ Excel file saved: {output_path}")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")

def create_detailed_dataframe(results):
    """
    Create a detailed pandas DataFrame with all results
    
    Args:
        results (list): List of result data dictionaries
        
    Returns:
        pd.DataFrame: DataFrame with all result data
    """
    # Flatten the results data for DataFrame
    flattened_data = []
    
    for result in results:
        base_data = {
            'PDF File': result.get('pdf_file', ''),
            'USN': result.get('usn', ''),
            'Student Name': result.get('student_name', ''),
            'Semester': result.get('semester', ''),
            'Result': result.get('result', ''),
            'SGPA': result.get('sgpa', ''),
            'CGPA': result.get('cgpa', ''),
            'Total Marks': result.get('total_marks', ''),
            'Credits Registered': result.get('credits_registered', ''),
            'Credits Earned': result.get('credits_earned', '')
        }
        
        # If no subjects, add base data once
        if not result.get('subjects'):
            flattened_data.append(base_data)
        else:
            # Add a row for each subject
            for subject in result.get('subjects', []):
                subject_data = base_data.copy()
                subject_data.update({
                    'Subject Code': subject.get('subject_code', ''),
                    'Subject Name': subject.get('subject_name', ''),
                    'Internal': subject.get('internal', ''),
                    'External': subject.get('external', ''),
                    'Total': subject.get('total', ''),
                    'Subject Result': subject.get('result', '')
                })
                flattened_data.append(subject_data)
    
    return pd.DataFrame(flattened_data)

def create_subject_wise_dataframe(results):
    """
    Create a subject-wise DataFrame with subjects as columns
    
    Args:
        results (list): List of result data dictionaries
        
    Returns:
        pd.DataFrame: DataFrame with subjects as columns
    """
    # Collect all unique subject codes (not names)
    all_subjects = set()
    for result in results:
        for subject in result.get('subjects', []):
            all_subjects.add(subject.get('subject_code', '').strip())
    
    # Sort subject codes for consistent column ordering
    subject_columns = sorted(list(all_subjects))
    
    # Create rows for each student
    rows = []
    for result in results:
        # Clean up the student name
        student_name = result.get('student_name', '')
        if '\n' in student_name:
            student_name = student_name.split('\n')[0].strip()
        
        row = {
            'USN': result.get('usn', ''),
            'Name': student_name,  # Changed from 'NAME' to 'Name'
        }
        
        # Add subject marks
        for subject_code in subject_columns:
            row[subject_code] = ''  # Initialize with empty string
        
        # Fill in actual subject marks
        for subject in result.get('subjects', []):
            subject_code = subject.get('subject_code', '').strip()
            if subject_code in row:
                row[subject_code] = subject.get('total', '')
        
        # Add CGPA and SGPA from result data
        row['CGPA'] = result.get('cgpa', '')
        row['SGPA'] = result.get('sgpa', '')
        
        # Determine overall result
        subject_results = [subject.get('result', '').upper() for subject in result.get('subjects', [])]
        if all(res == 'P' for res in subject_results) and subject_results:
            row['Result'] = 'P'  # Changed from 'RESULT' to 'Result'
        elif any(res == 'F' for res in subject_results):
            # Find the first failed subject
            failed_subject = None
            for subject in result.get('subjects', []):
                if subject.get('result', '').upper() == 'F':
                    failed_subject = subject.get('subject_code', '').strip()
                    break
            row['Result'] = f'F in {failed_subject}' if failed_subject else 'F'  # Changed from 'RESULT' to 'Result'
        else:
            row['Result'] = 'N/A'  # Changed from 'RESULT' to 'Result'
        
        rows.append(row)
    
    # Define column order to match exact requirements: USN, Name, subject codes, Result, CGPA, SGPA
    columns = ['USN', 'Name'] + subject_columns + ['Result', 'CGPA', 'SGPA']
    
    return pd.DataFrame(rows, columns=columns)

def main():
    """
    Main function to process PDF results and convert to Excel
    """
    print("🚀 VTU PDF Results to Excel Converter")
    print("=" * 50)
    
    # Get the project root directory
    project_root = Path(__file__).parent
    pdf_results_dir = project_root / "pdf_results"
    output_dir = project_root / "excel_results"
    output_dir.mkdir(exist_ok=True)
    
    print(f"📂 PDF Results Directory: {pdf_results_dir}")
    print(f"📂 Output Directory: {output_dir}")
    
    # Process all PDF files
    results = process_pdf_batch(pdf_results_dir)
    
    if not results:
        print("❌ No PDF results found to process")
        return
    
    print(f"✅ Processed {len(results)} PDF files")
    
    # Create only the subject-wise DataFrame with the exact format requested
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    subject_wise_df = create_subject_wise_dataframe(results)
    output_excel_path = output_dir / f"vtu_results_{timestamp}.xlsx"
    subject_wise_df.to_excel(output_excel_path, index=False)
    print(f"📊 Excel file saved: {output_excel_path}")
    
    # Also display the column structure for verification
    print(f"📋 Columns in output: {list(subject_wise_df.columns)}")
    
    print("\n🎉 Conversion completed successfully!")
    print(f"📁 Output file saved in: {output_dir}")

if __name__ == "__main__":
    main()